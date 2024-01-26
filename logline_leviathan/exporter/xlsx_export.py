from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from logline_leviathan.exporter.export_constructor import generate_dataframe


def generate_xlsx_file(output_file_path, db_session, checkboxes, context_selection, only_crossmatches):
    # Fetch data using the new DataFrame constructor
    df = generate_dataframe(db_session, checkboxes, context_selection, only_crossmatches)
    
    # Reorder columns
    # Check if 'Sources' or 'Source File' and 'Line Number' columns are in the DataFrame
    if 'Sources' in df.columns:
        df = df[["Entity Type", "Entity", "Occurrences", "Timestamp", "Sources", "Context"]]
    elif 'Source File' in df.columns and 'Line Number' in df.columns:
        df = df[["Entity Type", "Entity", "Occurrences", "Timestamp", "Source File", "Line Number", "Context"]]


    # Create a new Workbook
    workbook = Workbook()

    # Remove the default sheet created by openpyxl
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    # Iterate over unique entity types
    for entity_type in df['Entity Type'].unique():
        # Filter the DataFrame for the current entity type
        df_filtered = df[df['Entity Type'] == entity_type]

        # Add a new worksheet for each entity type
        sheet = workbook.create_sheet(title=entity_type)

        # Set column width and enable text wrapping
        for column_cells in sheet.columns:
            sheet.column_dimensions[column_cells[0].column_letter].width = 20  # Standard width for all columns
            for cell in column_cells:
                cell.alignment = Alignment(wrap_text=True)

        # Write header
        for col_num, column_title in enumerate(df_filtered.columns, start=1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = column_title
            cell.alignment = Alignment(wrap_text=True)

        # Write data rows
        for row_idx, row in enumerate(df_filtered.itertuples(), start=2):  # Start from row 2
            for col_idx, value in enumerate(row[1:], start=1):  # Skip index from row tuple
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(wrap_text=True)

    # Save the workbook
    workbook.save(output_file_path)
