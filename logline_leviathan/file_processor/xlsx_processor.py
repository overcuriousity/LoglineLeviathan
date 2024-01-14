import logging
import re
from openpyxl import load_workbook
from datetime import datetime
from logline_leviathan.database.database_manager import EntityTypesTable
from logline_leviathan.file_processor.file_database_ops import handle_file_metadata, handle_individual_entity, handle_distinct_entity, handle_context_snippet

def read_xlsx_content(file_path):
    try:
        workbook = load_workbook(filename=file_path)
        return workbook
    except Exception as e:
        logging.error(f"Error reading XLSX file {file_path}: {e}")
        return None

def get_line_numbers_from_pos(content, start_pos, end_pos):
    # For XLSX, the line number is the row number in the current sheet
    start_line = end_line = 0
    current_pos = 0
    for i, line in enumerate(content):
        current_pos += len(line)
        if start_pos < current_pos:
            start_line = i
            break
    for i, line in enumerate(content[start_line:], start=start_line):
        current_pos += len(line)
        if end_pos <= current_pos:
            end_line = i
            break
    return start_line, end_line


def find_timestamp_before_match(content, match_start_pos):
    search_content = content[:match_start_pos]
    timestamp_patterns = [
        (r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}', '%Y-%m-%d %H:%M:%S'),  # ISO 8601 Extended
        (r'\d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2}', '%Y/%m/%d %H:%M:%S'),  # ISO 8601 with slashes
        (r'\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2}', '%d/%m/%Y %H:%M:%S'),  # European Date Format
        (r'\d{2}-\d{2}-\d{4} \d{2}:\d{2}:\d{2}', '%m-%d-%Y %H:%M:%S'),  # US Date Format
        (r'\d{8}_\d{6}', '%Y%m%d_%H%M%S'),                             # Compact Format
        (r'\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}', '%Y-%m-%dT%H:%M:%S'),  # ISO 8601 Basic
        (r'\d{2}\.\d{2}\.\d{4} \d{2}:\d{2}:\d{2}', '%d.%m.%Y %H:%M:%S'),# German Date Format
        (r'\d{4}\d{2}\d{2} \d{2}:\d{2}:\d{2}', '%Y%m%d %H:%M:%S'),      # Basic Format without Separators
        (r'\d{1,2}-[A-Za-z]{3}-\d{4} \d{2}:\d{2}:\d{2}', '%d-%b-%Y %H:%M:%S'), # English Date Format with Month Name
        (r'(?:19|20)\d{10}', '%Y%m%d%H%M'),                             # Compact Numeric Format
        # Add more patterns as needed
    ]
    for pattern, date_format in timestamp_patterns:
        for timestamp_match in reversed(list(re.finditer(pattern, search_content))):
            try:
                # Convert the matched timestamp to the standardized format
                matched_timestamp = datetime.strptime(timestamp_match.group(), date_format)
                return matched_timestamp.strftime('%Y-%m-%d %H:%M:%S')
            except ValueError:
                continue  # If conversion fails, continue to the next pattern
    return None


def process_xlsx_file(file_path, file_mimetype, thread_instance, db_session, abort_flag):
    try:
        logging.info(f"Starting processing of XLSX file: {file_path}")
        workbook = read_xlsx_content(file_path)
        regex_patterns = db_session.query(EntityTypesTable).filter(EntityTypesTable.regex_pattern != None, EntityTypesTable.regex_pattern != '').all()

        if workbook is None:
            return 0

        entity_count = 0

        for sheet in workbook:
            sheet_name = sheet.title
            file_metadata = handle_file_metadata(db_session, file_path, file_mimetype, sheet_name=sheet_name)

            if abort_flag():
                logging.info("Processing aborted.")
                return entity_count

            content = [' '.join([str(cell.value) if cell.value is not None else ' ' for cell in row]) for row in sheet.iter_rows()]
            full_content = '\n'.join(content)
            for regex in regex_patterns:
                if not regex.regex_pattern.strip():
                    continue
                for match in re.finditer(regex.regex_pattern, full_content):
                    match_text = match.group()
                    if not match_text.strip():
                        continue

                    timestamp = find_timestamp_before_match(full_content, match.start())
                    match_start_line, match_end_line = get_line_numbers_from_pos(content, match.start(), match.end())

                    entity = handle_distinct_entity(db_session, match_text, regex.entity_type_id)
                    individual_entity = handle_individual_entity(db_session, entity, file_metadata, match_start_line, timestamp, regex.entity_type_id, abort_flag, thread_instance)

                    if individual_entity:
                        handle_context_snippet(db_session, individual_entity, content, match_start_line, match_end_line)
                        entity_count += 1

            thread_instance.update_status.emit(f"   Finished processing sheet: {sheet_name}")

        logging.info(f"   Finished processing XLSX file: {file_path}")
        return entity_count
    except Exception as e:
        db_session.rollback()
        logging.error(f"Error processing XLSX file {file_path}: {e}")
        return 0

